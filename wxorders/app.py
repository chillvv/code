import os
import threading
import time
from random import uniform
from typing import List, Dict, Any
from flask import Flask, request, send_from_directory, jsonify

app = Flask(__name__, static_folder='static', static_url_path='/static')

# Global control for stopping send
stop_flag = threading.Event()


def _to_str(val: Any) -> str:
	if val is None:
		return ''
	# Avoid scientific notation and trailing .0 for integers
	if isinstance(val, (int,)):
		return str(val)
	if isinstance(val, float):
		if val.is_integer():
			return str(int(val))
			# Keep decimals if truly decimal
		return ('%f' % val).rstrip('0').rstrip('.')
	return str(val)


def load_excel(file_path: str) -> List[Dict[str, Any]]:
	"""Load Excel (xls/xlsx) into list of dicts with headers from first row.
	Adds _row_index as physical row number (starting at 1 for first data row)."""
	rows: List[Dict[str, Any]] = []
	lower = file_path.lower()
	if lower.endswith('.xlsx') or lower.endswith('.xlsm'):
		from openpyxl import load_workbook
		wb = load_workbook(file_path, read_only=True, data_only=True)
		ws = wb.active
		headers = []
		for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
			if i == 1:
				headers = [(_to_str(c)).strip() for c in row]
				continue
			data = {}
			for j, h in enumerate(headers):
				val = row[j] if j < len(row) else None
				data[h] = _to_str(val)
			data['_row_index'] = i - 1
			rows.append(data)
	elif lower.endswith('.xls'):
		import xlrd
		book = xlrd.open_workbook(file_path)
		sheet = book.sheet_by_index(0)
		headers = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]
		for r in range(1, sheet.nrows):
			data = {}
			for c in range(sheet.ncols):
				data[headers[c]] = _to_str(sheet.cell_value(r, c))
			data['_row_index'] = r
			rows.append(data)
	else:
		raise ValueError('Unsupported file extension')
	return rows


GOODS_CANDS = ['商品信息', '商品', '商品名称']
PAY_CANDS = ['支付状态', '支付', '付款状态']
ORDER_CANDS = ['订单状态', '状态']
ADDRESS_CANDS = ['收货地址', '地址', '收货信息']
NOTE_CANDS = ['用户备注', '买家留言', '订单备注', '备注']


def pick(row: Dict, candidates: List[str]) -> str:
	for name in candidates:
		if name in row and str(row.get(name, '')).strip() != '':
			return str(row.get(name, '')).strip()
	return ''


def classify_and_filter(rows: List[Dict]):
	lunch_key = '明日午餐 x1'
	dinner_key = '明日晚餐 x1'

	def is_paid(status: str) -> bool:
		return status.strip() == '已支付'

	def is_excluded(order_status: str, pay_status: str) -> bool:
		s = order_status.strip()
		p = pay_status.strip()
		if p == '未支付':
			return True
		if s in ('已取消', '用户申请退款'):
			return True
		if p == '已退款':
			return True
		return False

	lunch, dinner = [], []
	for row in rows:
		goods = pick(row, GOODS_CANDS)
		pay = pick(row, PAY_CANDS)
		order_status = pick(row, ORDER_CANDS)
		if not is_paid(pay):
			continue
		if is_excluded(order_status, pay):
			continue
		if goods == lunch_key:
			lunch.append(row)
		elif goods == dinner_key:
			dinner.append(row)
	# Reverse by physical row index descending
	lunch_sorted = sorted(lunch, key=lambda r: int(r.get('_row_index', 0)), reverse=True)
	dinner_sorted = sorted(dinner, key=lambda r: int(r.get('_row_index', 0)), reverse=True)
	return lunch_sorted, dinner_sorted


def split_address(addr: str) -> str:
	# Expect format: 姓名 - 电话 - 详细地址, keep original symbols
	text = str(addr)
	parts = [p.strip() for p in text.split('-')]
	if len(parts) >= 3:
		name = parts[0]
		phone = parts[1]
		detail = '-'.join(parts[2:]).strip()
		return f"{name} - {phone} - {detail}"
	return text.strip()


def extract_user_note(note_field: str) -> str:
	note = str(note_field).strip()
	if note:
		return f"（用户备注：{note}）"
	return ''


def format_block(title: str, rows: List[Dict], start_no: int) -> str:
	lines: List[str] = []
	lines.append(title)
	current = start_no
	for r in rows:
		addr = split_address(pick(r, ADDRESS_CANDS))
		note = extract_user_note(pick(r, NOTE_CANDS))
		lines.append(str(current))
		lines.append(addr)
		if note:
			lines.append(note)
		current += 1
	return '\n'.join(lines)


@app.route('/')
def index():
	return send_from_directory(app.static_folder, 'index.html')


@app.post('/api/preview')
def api_preview():
	# Receive file via form-data 'file', lunchStart, dinnerStart
	f = request.files.get('file')
	lunch_start = int(request.form.get('lunchStart', '1'))
	dinner_start = int(request.form.get('dinnerStart', '1'))
	if not f:
		return jsonify({'error': 'no file'}), 400
	tmp_dir = os.path.join('static', 'uploads')
	os.makedirs(tmp_dir, exist_ok=True)
	tmp_path = os.path.join(tmp_dir, f'preview{os.path.splitext(f.filename or "")[1] or ".xlsx"}')
	f.save(tmp_path)
	rows = load_excel(tmp_path)
	lunch, dinner = classify_and_filter(rows)
	lunch_title = f"### 一、午餐（商品信息：明日午餐 x1，编号从{lunch_start}开始）"
	dinner_title = f"### 二、晚餐（商品信息：明日晚餐 x1，编号从{dinner_start}开始）"
	lunch_text = format_block(lunch_title, lunch, lunch_start)
	dinner_text = format_block(dinner_title, dinner, dinner_start)
	return jsonify({'lunch': lunch_text, 'dinner': dinner_text, 'counts': {'lunch': len(lunch), 'dinner': len(dinner)}})


# Windows-only: wxauto sender
try:
	from wxauto import WeChat  # type: ignore
	import win32gui  # type: ignore
	WX_AVAILABLE = True
except Exception:
	WX_AVAILABLE = False


def send_to_wechat_group(group_name: str, lines: List[str], delay_bounds=(1.0, 1.5)):
	if not WX_AVAILABLE:
		raise RuntimeError('WeChat automation not available on this system')
	wc = WeChat()
	wc.Connect()
	wc.ChatWith(group_name)
	for line in lines:
		if stop_flag.is_set():
			break
		wc.SendMsg(line)
		time.sleep(uniform(*delay_bounds))


@app.post('/api/send')
def api_send():
	# body: { textLunch, textDinner, target: 'lunch'|'dinner'|'both', test: bool }
	data = request.get_json(force=True)
	target = data.get('target', 'both')
	text_lunch = data.get('textLunch', '')
	text_dinner = data.get('textDinner', '')
	test_mode = bool(data.get('test', True))

	groups = {
		'lunch': '简知午餐群',
		'dinner': '简知晚餐群',
		'test': '末'
	}

	stop_flag.clear()

	def job():
		try:
			if test_mode:
				if text_lunch.strip():
					send_to_wechat_group(groups['test'], text_lunch.split('\n'))
				if text_dinner.strip():
					send_to_wechat_group(groups['test'], text_dinner.split('\n'))
			else:
				if target in ('lunch', 'both') and text_lunch.strip():
					send_to_wechat_group(groups['lunch'], text_lunch.split('\n'))
				if target in ('dinner', 'both') and text_dinner.strip():
					send_to_wechat_group(groups['dinner'], text_dinner.split('\n'))
		except Exception as e:
			app.logger.exception('send failed: %s', e)

	t = threading.Thread(target=job, daemon=True)
	t.start()
	return jsonify({'status': 'started'})


@app.post('/api/stop')
def api_stop():
	stop_flag.set()
	return jsonify({'status': 'stopping'})


if __name__ == '__main__':
	port = int(os.environ.get('PORT', '5000'))
	app.run(host='0.0.0.0', port=port, debug=True)