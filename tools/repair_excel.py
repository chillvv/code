import sys
import os
from typing import Optional

def repair_to_xlsx(input_path: str, output_path: Optional[str] = None) -> str:
    if not os.path.exists(input_path):
        raise FileNotFoundError(input_path)
    root, _ = os.path.splitext(input_path)
    output = output_path or (root + "_repaired.xlsx")
    # First try openpyxl direct load/save
    try:
        from openpyxl import load_workbook, Workbook
        wb = load_workbook(filename=input_path, read_only=False, data_only=True)
        wb.save(output)
        return output
    except Exception:
        pass
    # Try xlrd read and write via openpyxl
    try:
        import xlrd  # type: ignore
        from openpyxl import Workbook
        book = xlrd.open_workbook(input_path, formatting_info=False, on_demand=True)
        out_wb = Workbook()
        # Remove the default sheet
        out_wb.remove(out_wb.active)
        for sheet_name in book.sheet_names():
            sh = book.sheet_by_name(sheet_name)
            ws = out_wb.create_sheet(title=sheet_name[:31])
            for r in range(sh.nrows):
                row_vals = sh.row_values(r)
                ws.append(row_vals)
        out_wb.save(output)
        return output
    except Exception as e:
        raise RuntimeError(f"无法修复该文件：{e}")

def main():
    if len(sys.argv) < 2:
        print("Usage: python -m tools.repair_excel <input-path> [output-path]")
        sys.exit(1)
    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) >= 3 else None
    out = repair_to_xlsx(input_path, output_path)
    print(out)

if __name__ == '__main__':
    main()

