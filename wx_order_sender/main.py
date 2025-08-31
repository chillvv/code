import os
import sys
import platform
import threading
import asyncio
import time
import random
import traceback
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import pandas as pd

from PyQt5 import QtCore, QtGui, QtWidgets
from qasync import QEventLoop, asyncSlot
import aiohttp
from openpyxl import load_workbook  # robust XLSX reader


BOT_NOT_READY_MSG = "Bot 未连接，请先设置模式/Token 并点击‘连接’"


REQUIRED_COLUMNS = ["商品信息", "支付状态", "订单状态", "收货地址", "用户备注"]


@dataclass
class ColumnMapping:
    product_info: str
    pay_status: str
    order_status: str
    address: str
    user_note: str


class DropArea(QtWidgets.QFrame):
    fileDropped = QtCore.pyqtSignal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.setStyleSheet("""
            QFrame {
                border: 2px dashed #999;
                border-radius: 8px;
                background: #fafafa;
            }
        """)
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        self.label = QtWidgets.QLabel("将 Excel/CSV 文件拖拽到此处，或点击下方按钮选择文件")
        self.label.setAlignment(QtCore.Qt.AlignCenter)
        self.label.setStyleSheet("color:#555;font-size:14px;")
        layout.addWidget(self.label)

    def dragEnterEvent(self, event: QtGui.QDragEnterEvent) -> None:
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event: QtGui.QDropEvent) -> None:
        urls = event.mimeData().urls()
        if not urls:
            return
        local_path = urls[0].toLocalFile()
        if local_path:
            self.fileDropped.emit(local_path)


def detect_csv_encoding(file_path: str) -> str:
    try:
        import chardet  # type: ignore
    except Exception:
        return "utf-8"
    with open(file_path, "rb") as f:
        raw = f.read(4096)
    result = chardet.detect(raw)
    enc = result.get("encoding") or "utf-8"
    return enc


def load_dataframe(file_path: str, sheet_name: Optional[str] = None) -> Tuple[pd.DataFrame, List[str]]:
    """Load Excel/CSV broadly: supports xlsx/xlsm/xltx/xltm/xls/xlsb/ods/csv/txt.
    Fallback through multiple engines where possible.
    """
    ext = os.path.splitext(file_path)[1].lower()

    def try_excel_with_engines(engines: List[str]) -> Tuple[pd.DataFrame, List[str]]:
        last_err: Optional[Exception] = None
        for eng in engines:
            try:
                xls = pd.ExcelFile(file_path, engine=eng)
                sheets = xls.sheet_names
                df = pd.read_excel(xls, sheet_name=sheet_name or sheets[0])
                return df, sheets
            except Exception as e:
                last_err = e
                continue
        if last_err:
            raise last_err
        raise RuntimeError("无法读取 Excel 文件")

    # Excel formats
    if ext in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        # Prefer using openpyxl directly for robustness with partially damaged files
        return load_xlsx_via_openpyxl(file_path, sheet_name)
    if ext == ".xls":
        # Prefer pyexcel-xls for legacy binary Excel
        try:
            return load_xls_via_pyexcel(file_path, sheet_name)
        except Exception:
            # As a fallback, attempt pandas engine chain (may fail with xlrd>=2)
            return try_excel_with_engines(["xlrd"])  # legacy Excel
    if ext == ".xlsb":
        return try_excel_with_engines(["pyxlsb"])  # binary Excel
    if ext == ".ods":
        return try_excel_with_engines(["odf"])  # OpenDocument Spreadsheet

    # CSV/TXT with encoding and delimiter sniffing
    if ext in [".csv", ".txt"]:
        enc = detect_csv_encoding(file_path)
        try:
            # sep=None uses Python engine's sniffing
            df = pd.read_csv(file_path, encoding=enc, sep=None, engine="python")
        except Exception:
            # Fallback to utf-8 and common separators
            try:
                df = pd.read_csv(file_path, encoding="utf-8", sep=None, engine="python", errors="ignore")
            except Exception:
                df = pd.read_csv(file_path, encoding="utf-8", sep=",", engine="python", errors="ignore")
        return df, ["CSV"]

    # Unknown extension: still try Excel engines in broad order
    try:
        return try_excel_with_engines(["openpyxl", "pyxlsb", "odf"]) 
    except Exception:
        pass
    # As a last resort, try reading as CSV
    try:
        enc = detect_csv_encoding(file_path)
        df = pd.read_csv(file_path, encoding=enc, sep=None, engine="python")
        return df, ["CSV"]
    except Exception:
        raise ValueError("不受支持的文件类型，请提供 Excel 或 CSV 文件")


def _guess_header_and_data(rows: List[List[object]]) -> Tuple[List[str], List[List[object]]]:
    # Find the row within the first 10 lines with the most non-empty cells as header
    max_nonempty = -1
    header_idx = 0
    for i in range(min(10, len(rows))):
        nonempty = sum(1 for v in rows[i] if v not in (None, ""))
        if nonempty > max_nonempty:
            max_nonempty = nonempty
            header_idx = i
    headers_raw = rows[header_idx] if rows else []
    headers = [str(h).strip() if h is not None else f"列{i+1}" for i, h in enumerate(headers_raw)]
    data_rows = rows[header_idx + 1 :] if header_idx + 1 <= len(rows) else []
    # Normalize row lengths to headers
    width = len(headers)
    normalized = []
    for r in data_rows:
        r = list(r)
        if len(r) < width:
            r = r + [None] * (width - len(r))
        elif len(r) > width:
            r = r[:width]
        normalized.append(r)
    return headers, normalized


def load_xlsx_via_openpyxl(file_path: str, sheet_name: Optional[str]) -> Tuple[pd.DataFrame, List[str]]:
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    sheets = wb.sheetnames
    name = sheet_name or (sheets[0] if sheets else None)
    if name is None:
        return pd.DataFrame(), []
    ws = wb[name]
    rows: List[List[object]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    headers, data_rows = _guess_header_and_data(rows)
    df = pd.DataFrame(data_rows, columns=headers)
    return df, sheets


def load_xls_via_pyexcel(file_path: str, sheet_name: Optional[str]) -> Tuple[pd.DataFrame, List[str]]:
    # Requires: pyexcel-xls, pyexcel-io
    from pyexcel_xls import get_data  # type: ignore
    data_dict = get_data(file_path)
    sheets = list(data_dict.keys())
    name = sheet_name or (sheets[0] if sheets else None)
    if name is None:
        return pd.DataFrame(), []
    rows = data_dict[name]
    headers, data_rows = _guess_header_and_data(rows)
    df = pd.DataFrame(data_rows, columns=headers)
    return df, sheets


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    # Flatten MultiIndex and normalize to string headers
    def _to_str(col) -> str:
        if isinstance(col, tuple):
            return " ".join([str(x).strip() for x in col])
        return str(col).strip()
    df.columns = [_to_str(c) for c in df.columns]
    return df


def infer_default_mapping(df: pd.DataFrame) -> ColumnMapping:
    cols = list(df.columns)
    mapping: Dict[str, str] = {}
    for target in REQUIRED_COLUMNS:
        if target in cols:
            mapping[target] = target
        else:
            # try fuzzy match by removing spaces
            target_key = target.replace(" ", "")
            found = None
            for c in cols:
                if str(c).replace(" ", "") == target_key:
                    found = c
                    break
            mapping[target] = found or (cols[0] if cols else target)
    return ColumnMapping(
        product_info=mapping["商品信息"],
        pay_status=mapping["支付状态"],
        order_status=mapping["订单状态"],
        address=mapping["收货地址"],
        user_note=mapping["用户备注"],
    )


def split_address(name_phone_addr: str) -> str:
    if not isinstance(name_phone_addr, str):
        name_phone_addr = str(name_phone_addr) if pd.notna(name_phone_addr) else ""
    text = name_phone_addr.strip()
    if not text:
        return " -  - "

    # Prefer strict split by hyphen with spaces
    parts = [p.strip() for p in re_split_hyphen(text)]
    if len(parts) >= 3:
        name = parts[0]
        phone = parts[1]
        address = " - ".join(parts[2:])
        return f"{name} - {phone} - {address}"

    # Fallback: detect phone number and split around it
    phone_match = find_phone_like(text)
    if phone_match:
        start, end = phone_match.span()
        name = text[:start].strip(" -")
        phone = text[start:end]
        address = text[end:].strip(" -")
        return f"{name} - {phone} - {address}" if name or address else f" - {phone} - "

    # Last resort: keep as address only
    return f" -  - {text}"


def re_split_hyphen(text: str) -> List[str]:
    import re
    # Split on hyphen-like characters with optional spaces around
    return re.split(r"\s*[-\-\–\—\－]\s*", text)


def find_phone_like(text: str):
    import re
    # Mainland China phones often 11 digits, but keep flexible 6-20
    pattern = re.compile(r"(\d[\d\s-]{5,19}\d)")
    return pattern.search(text)


def filter_and_order(df: pd.DataFrame, mapping: ColumnMapping) -> Tuple[pd.DataFrame, pd.DataFrame]:
    df = df.copy()
    # Keep a physical row index to honor original order
    df["__row__"] = range(1, len(df) + 1)

    # Normalize NaNs to empty strings for comparisons
    for col in [mapping.product_info, mapping.pay_status, mapping.order_status, mapping.address, mapping.user_note]:
        if col in df.columns:
            df[col] = df[col].fillna("")

    def is_paid(row) -> bool:
        return str(row[mapping.pay_status]).strip() == "已支付"

    def is_excluded(row) -> bool:
        status = str(row[mapping.order_status]).strip()
        pay = str(row[mapping.pay_status]).strip()
        if pay == "未支付":
            return True
        if status in ("已取消", "用户申请退款"):
            return True
        if pay == "已退款":
            return True
        return False

    # Apply filters
    mask_paid = df.apply(is_paid, axis=1)
    mask_excluded = df.apply(is_excluded, axis=1)
    effective = df[mask_paid & (~mask_excluded)].copy()

    # Split categories by product info exact match
    lunch = effective[effective[mapping.product_info].astype(str).str.strip() == "明日午餐 x1"].copy()
    dinner = effective[effective[mapping.product_info].astype(str).str.strip() == "明日晚餐 x1"].copy()

    # Order by physical row from bottom to top: larger __row__ first
    lunch = lunch.sort_values("__row__", ascending=False)
    dinner = dinner.sort_values("__row__", ascending=False)

    return lunch, dinner


def build_output(df: pd.DataFrame, mapping: ColumnMapping, start_number: int, title_prefix: str, product_label: str) -> str:
    lines: List[str] = []
    lines.append(f"### {title_prefix}（商品信息：{product_label}，编号从{start_number}开始）")
    current = start_number
    for _, row in df.iterrows():
        address_formatted = split_address(str(row.get(mapping.address, "")))
        user_note = str(row.get(mapping.user_note, ""))
        lines.append(str(current))
        lines.append(address_formatted)
        if user_note.strip():
            lines.append(f"（用户备注：{user_note}）")
        current += 1
    return "\n".join(lines)


class SendController(QtCore.QObject):
    progressed = QtCore.pyqtSignal(str)
    finished_ok = QtCore.pyqtSignal()
    failed = QtCore.pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self._stop = asyncio.Event()
        self._session: Optional[aiohttp.ClientSession] = None

    async def ensure_session(self):
        if self._session is None or self._session.closed:
            self._session = aiohttp.ClientSession()
        return self._session

    def stop(self):
        if not self._stop.is_set():
            self._stop.set()

    async def _sleep_with_check(self, seconds: float) -> bool:
        try:
            await asyncio.wait_for(asyncio.sleep(seconds), timeout=seconds + 1)
            return not self._stop.is_set()
        except asyncio.CancelledError:
            return False

    async def send_texts(self, lunch_text: str, dinner_text: str, lunch_group: str, dinner_group: str, interval_min: float, interval_max: float, test_mode: bool):
        try:
            session = await self.ensure_session()
            targets = [(lunch_group, lunch_text), (dinner_group, dinner_text)]
            if test_mode:
                targets = [("末", lunch_text), ("末", dinner_text)]
            payload = {
                "items": [{"group": g, "text": t} for g, t in targets if str(g).strip()],
                "intervalMin": float(interval_min),
                "intervalMax": float(interval_max)
            }
            async with session.post("http://127.0.0.1:8788/send", json=payload, timeout=60) as resp:
                if resp.status != 200:
                    text = await resp.text()
                    raise RuntimeError(f"发送失败：{resp.status} {text}")
            self.finished_ok.emit()
        except Exception as e:
            self.failed.emit(str(e))


def split_message_chunks(text: str, max_len: int = 3500) -> List[str]:
    if len(text) <= max_len:
        return [text]
    parts = []
    current = []
    current_len = 0
    for line in text.splitlines():
        # +1 for the newline if not first line
        add_len = len(line) + (1 if current else 0)
        if current_len + add_len > max_len:
            parts.append("\n".join(current))
            current = [line]
            current_len = len(line)
        else:
            if current:
                current.append(line)
                current_len += len(line) + 1
            else:
                current = [line]
                current_len = len(line)
    if current:
        parts.append("\n".join(current))
    return parts


class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("简知订单整理发送器（Node Wechaty）")
        self.setMinimumSize(960, 680)

        self.df: Optional[pd.DataFrame] = None
        self.sheets: List[str] = []
        self.current_file: Optional[str] = None
        self.mapping: Optional[ColumnMapping] = None
        self.send_controller = SendController()
        self._send_task: Optional[asyncio.Task] = None

        self._init_ui()

    def _init_ui(self):
        central = QtWidgets.QWidget()
        self.setCentralWidget(central)
        root = QtWidgets.QVBoxLayout(central)
        root.setContentsMargins(16, 16, 16, 16)
        root.setSpacing(12)

        # Top: file controls
        drop = DropArea()
        drop.fileDropped.connect(self.on_file_dropped)
        root.addWidget(drop)

        file_row = QtWidgets.QHBoxLayout()
        self.file_label = QtWidgets.QLabel("未选择文件")
        self.file_label.setStyleSheet("color:#333;")
        file_btn = QtWidgets.QPushButton("选择文件…")
        file_btn.clicked.connect(self.on_pick_file)
        self.sheet_combo = QtWidgets.QComboBox()
        self.sheet_combo.currentIndexChanged.connect(self.on_sheet_changed)
        self.sheet_combo.setEnabled(False)
        file_row.addWidget(self.file_label, 1)
        file_row.addWidget(QtWidgets.QLabel("工作表："))
        file_row.addWidget(self.sheet_combo)
        file_row.addWidget(file_btn)
        root.addLayout(file_row)

        # Column mapping
        map_group = QtWidgets.QGroupBox("字段映射（自动识别，可手动调整）")
        grid = QtWidgets.QGridLayout(map_group)
        self.cmb_product = QtWidgets.QComboBox()
        self.cmb_pay = QtWidgets.QComboBox()
        self.cmb_status = QtWidgets.QComboBox()
        self.cmb_addr = QtWidgets.QComboBox()
        self.cmb_note = QtWidgets.QComboBox()
        for i, (label, widget) in enumerate([
            ("商品信息", self.cmb_product),
            ("支付状态", self.cmb_pay),
            ("订单状态", self.cmb_status),
            ("收货地址", self.cmb_addr),
            ("用户备注", self.cmb_note),
        ]):
            grid.addWidget(QtWidgets.QLabel(label+"："), i, 0)
            grid.addWidget(widget, i, 1)
        map_group.setEnabled(False)
        self.map_group = map_group
        root.addWidget(map_group)

        # Settings row
        settings_group = QtWidgets.QGroupBox("发送设置")
        form = QtWidgets.QGridLayout(settings_group)
        # Wechaty token controls
        self.token_edit = QtWidgets.QLineEdit()
        self.token_edit.setPlaceholderText("Wechaty Puppet Service Token")
        self.btn_connect = QtWidgets.QPushButton("连接")
        self.btn_disconnect = QtWidgets.QPushButton("断开")
        self.btn_disconnect.setEnabled(False)
        self.btn_connect.clicked.connect(self.on_connect)
        self.btn_disconnect.clicked.connect(self.on_disconnect)

        self.puppet_combo = QtWidgets.QComboBox()
        self.puppet_combo.addItems(["wechat(推荐)", "service(需Token)"])


        self.lunch_start = QtWidgets.QSpinBox()
        self.lunch_start.setRange(1, 100000)
        self.lunch_start.setValue(7)

        self.dinner_start = QtWidgets.QSpinBox()
        self.dinner_start.setRange(1, 100000)
        self.dinner_start.setValue(7)

        self.cmb_lunch_group = QtWidgets.QComboBox()
        self.cmb_lunch_group.setEditable(True)
        self.cmb_lunch_group.addItems(["简知午餐群", "末"]) 
        self.cmb_lunch_group.setCurrentText("简知午餐群")

        self.cmb_dinner_group = QtWidgets.QComboBox()
        self.cmb_dinner_group.setEditable(True)
        self.cmb_dinner_group.addItems(["简知晚餐群", "末"]) 
        self.cmb_dinner_group.setCurrentText("简知晚餐群")

        self.test_mode = QtWidgets.QCheckBox("测试模式（强制发送到：末）")
        self.test_mode.setChecked(False)

        self.min_interval = QtWidgets.QDoubleSpinBox()
        self.min_interval.setRange(0.1, 10.0)
        self.min_interval.setSingleStep(0.1)
        self.min_interval.setValue(1.0)

        self.max_interval = QtWidgets.QDoubleSpinBox()
        self.max_interval.setRange(0.1, 10.0)
        self.max_interval.setSingleStep(0.1)
        self.max_interval.setValue(1.5)

        form.addWidget(QtWidgets.QLabel("午餐起始编号："), 0, 0)
        form.addWidget(self.lunch_start, 0, 1)
        form.addWidget(QtWidgets.QLabel("晚餐起始编号："), 0, 2)
        form.addWidget(self.dinner_start, 0, 3)

        form.addWidget(QtWidgets.QLabel("模式："), 1, 0)
        form.addWidget(self.puppet_combo, 1, 1)
        form.addWidget(QtWidgets.QLabel("Token："), 1, 2)
        form.addWidget(self.token_edit, 1, 3)

        form.addWidget(QtWidgets.QLabel("午餐发送至："), 2, 0)
        form.addWidget(self.cmb_lunch_group, 2, 1)
        form.addWidget(QtWidgets.QLabel("晚餐发送至："), 2, 2)
        form.addWidget(self.cmb_dinner_group, 2, 3)

        form.addWidget(QtWidgets.QLabel("发送间隔（秒）："), 3, 0)
        interval_row = QtWidgets.QHBoxLayout()
        interval_row.addWidget(QtWidgets.QLabel("最小"))
        interval_row.addWidget(self.min_interval)
        interval_row.addSpacing(8)
        interval_row.addWidget(QtWidgets.QLabel("最大"))
        interval_row.addWidget(self.max_interval)
        interval_widget = QtWidgets.QWidget()
        interval_widget.setLayout(interval_row)
        form.addWidget(interval_widget, 3, 1, 1, 3)

        form.addWidget(self.test_mode, 4, 0, 1, 2)
        btns_widget = QtWidgets.QWidget()
        btns_layout = QtWidgets.QHBoxLayout(btns_widget)
        btns_layout.setContentsMargins(0,0,0,0)
        btns_layout.addWidget(self.btn_connect)
        btns_layout.addWidget(self.btn_disconnect)
        btns_layout.addStretch(1)
        form.addWidget(btns_widget, 4, 2, 1, 2)

        root.addWidget(settings_group)

        # Preview and actions
        action_row = QtWidgets.QHBoxLayout()
        self.btn_preview = QtWidgets.QPushButton("预览")
        self.btn_preview.clicked.connect(self.on_preview)
        self.btn_send = QtWidgets.QPushButton("发送")
        self.btn_send.clicked.connect(self.on_send)
        self.btn_stop = QtWidgets.QPushButton("停止（Ctrl+Shift+S）")
        self.btn_stop.clicked.connect(self.on_stop)
        action_row.addWidget(self.btn_preview)
        action_row.addWidget(self.btn_send)
        action_row.addWidget(self.btn_stop)
        action_row.addStretch(1)
        root.addLayout(action_row)

        self.preview = QtWidgets.QPlainTextEdit()
        self.preview.setReadOnly(True)
        self.preview.setPlaceholderText("预览区：点击‘预览’生成内容；点击‘发送’将按设置发送至对应群聊")
        root.addWidget(self.preview, 1)

        self.status = QtWidgets.QLabel("")
        self.status.setStyleSheet("color:#0a7;")
        root.addWidget(self.status)

        # Global hotkey (best effort; may require permissions on some systems)
        self._install_hotkey()

    def _install_hotkey(self):
        # Best-effort global hotkey using keyboard; safe if unavailable
        def worker():
            try:
                import keyboard  # type: ignore
                keyboard.add_hotkey("ctrl+shift+s", lambda: self.on_stop())
                # Keep thread alive
                while True:
                    time.sleep(1)
            except Exception:
                return
        t = threading.Thread(target=worker, daemon=True)
        t.start()

    # File interactions
    def on_pick_file(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "选择 Excel/CSV 文件", os.path.expanduser("~"), "表格文件 (*.xlsx *.xls *.csv *.txt)")
        if path:
            self.load_file(path)

    def on_file_dropped(self, path: str):
        self.load_file(path)

    def load_file(self, path: str):
        try:
            df, sheets = load_dataframe(path)
            df = normalize_columns(df)
            self.df = df
            self.sheets = sheets
            self.current_file = path
            self.file_label.setText(f"已加载：{os.path.basename(path)}")
            # Populate sheets
            self.sheet_combo.blockSignals(True)
            self.sheet_combo.clear()
            self.sheet_combo.addItems(sheets)
            self.sheet_combo.blockSignals(False)
            self.sheet_combo.setEnabled(len(sheets) > 1)

            # Populate mapping combos
            self.map_group.setEnabled(True)
            for cmb in [self.cmb_product, self.cmb_pay, self.cmb_status, self.cmb_addr, self.cmb_note]:
                cmb.clear()
                cmb.addItems([str(c) for c in self.df.columns])

            m = infer_default_mapping(self.df)
            self.cmb_product.setCurrentText(m.product_info)
            self.cmb_pay.setCurrentText(m.pay_status)
            self.cmb_status.setCurrentText(m.order_status)
            self.cmb_addr.setCurrentText(m.address)
            self.cmb_note.setCurrentText(m.user_note)

            self.status.setText("文件加载成功。可点击‘预览’查看整理结果。")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "加载失败", f"{e}")

    def on_sheet_changed(self, idx: int):
        if not self.current_file:
            return
        try:
            name = self.sheets[idx]
            df, _ = load_dataframe(self.current_file, sheet_name=name)
            df = normalize_columns(df)
            self.df = df
            for cmb in [self.cmb_product, self.cmb_pay, self.cmb_status, self.cmb_addr, self.cmb_note]:
                cmb.clear()
                cmb.addItems([str(c) for c in self.df.columns])
            m = infer_default_mapping(self.df)
            self.cmb_product.setCurrentText(m.product_info)
            self.cmb_pay.setCurrentText(m.pay_status)
            self.cmb_status.setCurrentText(m.order_status)
            self.cmb_addr.setCurrentText(m.address)
            self.cmb_note.setCurrentText(m.user_note)
            self.status.setText("工作表已切换。")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "切换失败", f"{e}")

    def get_mapping(self) -> Optional[ColumnMapping]:
        if not self.df is None:
            return ColumnMapping(
                product_info=self.cmb_product.currentText(),
                pay_status=self.cmb_pay.currentText(),
                order_status=self.cmb_status.currentText(),
                address=self.cmb_addr.currentText(),
                user_note=self.cmb_note.currentText(),
            )
        return None

    def _build_preview_texts(self) -> Tuple[str, str]:
        if self.df is None:
            raise RuntimeError("请先加载 Excel/CSV 文件")
        mapping = self.get_mapping()
        assert mapping is not None
        lunch_df, dinner_df = filter_and_order(self.df, mapping)
        lunch_text = build_output(lunch_df, mapping, self.lunch_start.value(), "一、午餐", "明日午餐 x1")
        dinner_text = build_output(dinner_df, mapping, self.dinner_start.value(), "二、晚餐", "明日晚餐 x1")
        return lunch_text, dinner_text

    def on_preview(self):
        try:
            lunch_text, dinner_text = self._build_preview_texts()
            combined = f"{lunch_text}\n\n{dinner_text}".strip()
            self.preview.setPlainText(combined)
            self.status.setText("预览已生成。请核对格式与编号。")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "预览失败", f"{e}\n\n{traceback.format_exc()}")

    @asyncSlot()
    async def on_connect(self):
        puppet = 'wechat' if self.puppet_combo.currentIndex() == 0 else 'service'
        token = self.token_edit.text().strip()
        if puppet == 'service' and not token:
            QtWidgets.QMessageBox.warning(self, "缺少 Token", "请选择 service 时需填写 Token")
            return
        try:
            session = await self.send_controller.ensure_session()
            async with session.post("http://127.0.0.1:8788/connect", json={"puppet": puppet, "token": token}, timeout=60) as resp:
                if resp.status != 200:
                    txt = await resp.text()
                    raise RuntimeError(txt)
            self.btn_connect.setEnabled(False)
            self.btn_disconnect.setEnabled(True)
            self.status.setText("已连接 Bot 服务")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "连接失败", str(e))

    @asyncSlot()
    async def on_disconnect(self):
        try:
            session = await self.send_controller.ensure_session()
            async with session.post("http://127.0.0.1:8788/disconnect", timeout=30) as resp:
                pass
            self.btn_connect.setEnabled(True)
            self.btn_disconnect.setEnabled(False)
            self.status.setText("已断开连接")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "断开失败", str(e))

    def on_send(self):
        try:
            lunch_text, dinner_text = self._build_preview_texts()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "发送失败", f"{e}")
            return

        # Validate interval
        mi = float(self.min_interval.value())
        ma = float(self.max_interval.value())
        if ma < mi:
            QtWidgets.QMessageBox.warning(self, "参数错误", "最大发送间隔不能小于最小发送间隔")
            return

        lunch_group = self.cmb_lunch_group.currentText().strip()
        dinner_group = self.cmb_dinner_group.currentText().strip()
        if not lunch_group and not dinner_group and not self.test_mode.isChecked():
            QtWidgets.QMessageBox.warning(self, "缺少群聊", "请至少设置一个群聊或开启测试模式")
            return

        if self._send_task and not self._send_task.done():
            QtWidgets.QMessageBox.information(self, "提示", "发送任务进行中…")
            return
        self.btn_send.setEnabled(False)
        self.send_controller._stop.clear()
        coro = self.send_controller.send_texts(
            lunch_text=lunch_text,
            dinner_text=dinner_text,
            lunch_group=lunch_group,
            dinner_group=dinner_group,
            interval_min=mi,
            interval_max=ma,
            test_mode=self.test_mode.isChecked(),
        )
        loop = asyncio.get_event_loop()
        self._send_task = loop.create_task(coro)
        self.send_controller.progressed.connect(self.on_progress)
        self.send_controller.finished_ok.connect(self.on_finished)
        self.send_controller.failed.connect(self.on_failed)
        self.status.setText("正在发送…")

    def on_stop(self):
        if self._send_task and not self._send_task.done():
            asyncio.create_task(self._post_stop())
            self.status.setText("停止指令已发送，将尽快停止。")

    async def _post_stop(self):
        try:
            session = await self.send_controller.ensure_session()
            async with session.post("http://127.0.0.1:8788/stop", timeout=10):
                pass
        except Exception:
            pass

    def on_progress(self, msg: str):
        self.status.setText(msg)

    def on_finished(self):
        self.btn_send.setEnabled(True)
        self.status.setText("发送完成。")

    def on_failed(self, err: str):
        self.btn_send.setEnabled(True)
        QtWidgets.QMessageBox.critical(self, "发送失败", err)


def main():
    # Ensure stdout uses UTF-8 to avoid mojibake
    if platform.system().lower() == "windows":
        try:
            import ctypes
            ctypes.windll.kernel32.SetConsoleOutputCP(65001)
        except Exception:
            pass
    app = QtWidgets.QApplication(sys.argv)
    loop = QEventLoop(app)
    asyncio.set_event_loop(loop)
    win = MainWindow()
    win.show()
    with loop:
        loop.run_forever()


if __name__ == "__main__":
    main()

