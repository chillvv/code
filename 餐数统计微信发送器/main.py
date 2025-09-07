#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
简知轻食餐数统计微信发送器
基于扣餐表统计今日用餐人员并发送个人微信消息
"""

import os
import sys
import time
import random
import threading
import platform
import traceback
import tempfile
from typing import List, Optional, Tuple, Dict
from datetime import datetime, date

import pandas as pd
from PyQt5 import QtCore, QtGui, QtWidgets


def detect_csv_encoding(file_path: str) -> str:
    """检测CSV文件编码"""
    try:
        import chardet
    except Exception:
        return "utf-8"
    with open(file_path, "rb") as f:
        raw = f.read(4096)
    result = chardet.detect(raw)
    return result.get("encoding") or "utf-8"


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """标准化列名"""
    df = df.copy()
    def _to_str(col) -> str:
        if isinstance(col, tuple):
            return " ".join([str(x).strip() for x in col])
        return str(col).strip()
    df.columns = [_to_str(c) for c in df.columns]
    return df


def load_excel_file(file_path: str) -> Tuple[pd.DataFrame, List[str]]:
    """加载Excel文件"""
    ext = os.path.splitext(file_path)[1].lower()
    
    if ext in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            sheets = wb.sheetnames
            
            # 查找扣餐表
            target_sheet = None
            for sheet in sheets:
                if "扣餐表" in sheet or "扣餐" in sheet:
                    target_sheet = sheet
                    break
            
            if not target_sheet:
                target_sheet = sheets[0]
            
            # 读取数据
            df = pd.read_excel(file_path, sheet_name=target_sheet)
            return df, [target_sheet]
            
        except Exception as e:
            raise RuntimeError(f"Excel文件读取失败: {e}")
    
    elif ext == ".csv":
        enc = detect_csv_encoding(file_path)
        try:
            df = pd.read_csv(file_path, encoding=enc, sep=None, engine="python")
        except Exception:
            df = pd.read_csv(file_path, encoding="utf-8", sep=",", engine="python", errors="ignore")
        return df, ["CSV"]
    
    else:
        raise RuntimeError(f"不支持的文件格式: {ext}")


def analyze_meal_data(df: pd.DataFrame, target_date: int) -> Tuple[List[Dict], str]:
    """分析餐数数据，返回今日用餐人员信息"""
    
    # 清理数据
    df_clean = df.dropna(subset=['会员姓名']).copy()
    df_clean = df_clean[df_clean['会员姓名'].astype(str) != 'nan'].copy()
    
    # 检查是否有目标日期列
    if target_date not in df_clean.columns:
        available_dates = [col for col in df_clean.columns if isinstance(col, int) and 1 <= col <= 31]
        raise RuntimeError(f"未找到{target_date}号的数据列。可用日期: {available_dates}")
    
    # 筛选今日用餐人员
    today_diners = df_clean[
        df_clean[target_date].notna() & 
        (df_clean[target_date].astype(str).str.strip() != '') &
        (df_clean[target_date].astype(str).str.strip() != 'nan')
    ].copy()
    
    # 构建发送列表
    messages_to_send = []
    stats_summary = f"今日({target_date}号)用餐统计:\n"
    stats_summary += f"用餐人数: {len(today_diners)}\n\n"
    
    for idx, row in today_diners.iterrows():
        name = str(row['会员姓名']).strip()
        phone = str(row['电话']) if pd.notna(row['电话']) else '无电话'
        
        # 计算餐数
        initial_meals = row['剩余餐数'] if pd.notna(row['剩余餐数']) else 0
        remaining_meals = row['剩余'] if pd.notna(row['剩余']) else 0
        used_meals = initial_meals - remaining_meals if pd.notna(initial_meals) and pd.notna(remaining_meals) else 0
        
        # 今天的用餐信息
        today_meal_info = str(row[target_date]) if pd.notna(row[target_date]) else ''
        
        # 处理负数情况（可能是充值了餐数）
        if used_meals < 0:
            used_meals = "计算中"
            display_used = "计算中"
        else:
            display_used = f"{int(used_meals)}"
        
        # 生成个人消息
        personal_message = f"""亲爱的{name}，您好！

今天您已用餐，餐数统计如下：
📊 今日用餐：已记录
🍽️ 本月已用餐：{display_used}次  
💰 剩余餐数：{int(remaining_meals) if isinstance(remaining_meals, (int, float)) else remaining_meals}次

感谢您选择简知轻食！祝您用餐愉快！😊"""
        
        messages_to_send.append({
            'name': name,
            'phone': phone,
            'message': personal_message,
            'today_meal': today_meal_info,
            'used_meals': display_used,
            'remaining_meals': int(remaining_meals) if isinstance(remaining_meals, (int, float)) else remaining_meals
        })
        
        # 添加到统计摘要
        stats_summary += f"• {name}: 已用{display_used}次, 剩余{int(remaining_meals) if isinstance(remaining_meals, (int, float)) else remaining_meals}次\n"
    
    return messages_to_send, stats_summary


class DropArea(QtWidgets.QFrame):
    """文件拖拽区域"""
    fileDropped = QtCore.pyqtSignal(str)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.setStyleSheet("""
            QFrame { border: 2px dashed #bbb; border-radius: 8px; background: #fafafa; }
        """)
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(16,16,16,16)
        label = QtWidgets.QLabel("将扣餐表Excel文件拖拽到此处，或点击下方按钮选择文件")
        label.setAlignment(QtCore.Qt.AlignCenter)
        label.setStyleSheet("color:#555;font-size:14px;")
        layout.addWidget(label)
    
    def dragEnterEvent(self, e: QtGui.QDragEnterEvent) -> None:
        e.acceptProposedAction() if e.mimeData().hasUrls() else e.ignore()
    
    def dropEvent(self, e: QtGui.QDropEvent) -> None:
        urls = e.mimeData().urls()
        if urls:
            self.fileDropped.emit(urls[0].toLocalFile())


class WeChatPersonalSender(QtCore.QObject):
    """微信个人消息发送器 - 增强版"""
    progressed = QtCore.pyqtSignal(str)
    finished = QtCore.pyqtSignal()
    failed = QtCore.pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self._stop = threading.Event()
        self.wechat_path: Optional[str] = None
        self.wechat_hwnd = None
        
        # 检查依赖
        self.HAS_WIN32 = False
        self.HAS_AUTO = False
        self.HAS_PYAUTOGUI = False
        
        try:
            import win32gui, win32con
            self.HAS_WIN32 = True
        except ImportError:
            pass
            
        try:
            import uiautomation
            self.HAS_AUTO = True
        except ImportError:
            pass
            
        try:
            import pyautogui
            import pyperclip
            self.HAS_PYAUTOGUI = True
        except ImportError:
            pass

    def stop(self):
        self._stop.set()

    def _sleep(self, seconds: float) -> bool:
        """可中断的睡眠"""
        end = time.time() + seconds
        while time.time() < end:
            if self._stop.is_set():
                return False
            time.sleep(0.05)
        return True

    def _ensure_wechat(self):
        """确保微信已启动"""
        import uiautomation as auto
        main = auto.WindowControl(searchDepth=1, ClassName="WeChatMainWndForPC")
        if not main.Exists(0.5):
            try:
                exe = self._get_wechat_exe_path()
                if exe:
                    os.startfile(exe)
            except Exception:
                pass
            # 等待微信启动
            start = time.time()
            while time.time() - start < 20:
                if main.Exists(0.5):
                    break
                time.sleep(0.5)
        if not main.Exists(0.5):
            raise RuntimeError("未找到微信窗口，请先登录微信")
        return main

    def _get_wechat_exe_path(self) -> Optional[str]:
        """获取微信可执行文件路径"""
        candidates = [
            'WeChat.exe',
            r'C:\Program Files (x86)\Tencent\WeChat\WeChat.exe',
            r'C:\Program Files\Tencent\WeChat\WeChat.exe',
            r'D:\Program Files (x86)\Tencent\WeChat\WeChat.exe',
            r'D:\Program Files\Tencent\WeChat\WeChat.exe',
        ]
        
        if self.wechat_path:
            candidates.insert(0, self.wechat_path)
        
        for p in candidates:
            try:
                if os.path.isabs(p) and os.path.exists(p):
                    return p
                if p.lower().endswith('.exe') and not os.path.isabs(p):
                    return p
            except Exception:
                continue
        return None

    def _search_and_enter_chat(self, contact_name: str) -> bool:
        """搜索并进入聊天窗口"""
        try:
            if not self.HAS_PYAUTOGUI:
                return False
                
            import pyautogui
            import pyperclip
            
            self.progressed.emit(f"🔍 搜索联系人: {contact_name}")
            
            # 多次尝试打开搜索框
            search_attempts = 0
            max_attempts = 3
            
            while search_attempts < max_attempts:
                try:
                    pyautogui.hotkey('ctrl', 'f')
                    time.sleep(0.5)
                    search_attempts += 1
                    
                    # 清空搜索框
                    pyautogui.hotkey('ctrl', 'a')
                    time.sleep(0.2)
                    pyautogui.press('delete')
                    time.sleep(0.2)
                    
                    # 复制联系人姓名到剪贴板并验证
                    pyperclip.copy(contact_name)
                    time.sleep(0.1)
                    
                    # 验证剪贴板内容
                    clipboard_content = pyperclip.paste()
                    if clipboard_content != contact_name:
                        self.progressed.emit(f"⚠️ 剪贴板验证失败，重试...")
                        continue
                    
                    # 粘贴联系人姓名
                    pyautogui.hotkey('ctrl', 'v')
                    time.sleep(0.5)
                    
                    # 按回车进入聊天
                    pyautogui.press('enter')
                    time.sleep(1.5)  # 等待聊天窗口加载
                    
                    self.progressed.emit(f"✅ 已进入与 {contact_name} 的聊天")
                    return True
                    
                except Exception as e:
                    self.progressed.emit(f"⚠️ 搜索尝试 {search_attempts} 失败: {e}")
                    if search_attempts < max_attempts:
                        time.sleep(0.5)
                        continue
                    else:
                        break
            
            return False
            
        except Exception as e:
            self.progressed.emit(f"❌ 搜索联系人失败: {e}")
            return False

    def _find_input_box_position(self):
        """查找微信输入框的实际位置"""
        try:
            # 方法1: 尝试使用控件识别（如果可用）
            position = self._find_input_by_control()
            if position:
                return position
            
            # 方法2: 基于窗口计算
            position = self._find_input_by_window_calc()
            if position:
                return position
                
            return None
            
        except Exception as e:
            self.progressed.emit(f"查找输入框位置失败: {e}")
            return None

    def _find_input_by_control(self):
        """通过控件识别查找输入框"""
        try:
            if not self.HAS_AUTO or not self.wechat_hwnd:
                return None
                
            import uiautomation as auto
            import win32gui
            
            # 获取微信窗口的控制对象
            main_window = auto.WindowControl(Handle=self.wechat_hwnd)
            if not main_window.Exists(0.5):
                return None
            
            # 查找所有编辑框控件
            edit_controls = main_window.EditControls()
            if not edit_controls:
                return None
            
            # 从后往前遍历，找到最适合的输入框
            for edit_control in reversed(edit_controls):
                try:
                    rect = edit_control.BoundingRectangle
                    if rect.width() > 100 and rect.height() > 20:  # 合理的输入框尺寸
                        center_x = rect.left + rect.width() // 2
                        center_y = rect.top + rect.height() // 2
                        self.progressed.emit(f"通过控件找到输入框: ({center_x}, {center_y})")
                        return (center_x, center_y)
                except Exception:
                    continue
            
            return None
            
        except Exception as e:
            self.progressed.emit(f"控件识别查找输入框失败: {e}")
            return None

    def _find_input_by_window_calc(self):
        """通过窗口计算估算输入框位置"""
        try:
            if not self.HAS_WIN32 or not self.wechat_hwnd:
                return None
                
            import win32gui
            
            # 获取微信窗口位置和尺寸
            rect = win32gui.GetWindowRect(self.wechat_hwnd)
            window_width = rect[2] - rect[0]
            window_height = rect[3] - rect[1]
            
            # 基于窗口尺寸动态计算输入框位置
            # 输入框通常在窗口底部，距离底部约60-100像素
            base_offset = 80
            if window_height > 800:
                y_offset = 100
            elif window_height > 600:
                y_offset = 85
            else:
                y_offset = 70
            
            # 计算输入框中心位置
            center_x = rect[0] + window_width // 2
            center_y = rect[3] - y_offset
            
            self.progressed.emit(f"通过窗口计算输入框位置: ({center_x}, {center_y})")
            return (center_x, center_y)
            
        except Exception as e:
            self.progressed.emit(f"窗口计算查找输入框失败: {e}")
            return None

    def _smart_click_input_area(self):
        """智能点击输入区域并验证"""
        try:
            if not self.HAS_PYAUTOGUI or not self.wechat_hwnd:
                return False
                
            import pyautogui
            import win32gui
            
            # 获取微信窗口位置
            rect = win32gui.GetWindowRect(self.wechat_hwnd)
            window_width = rect[2] - rect[0]
            
            # 尝试多个可能的输入区域位置
            base_y = rect[3] - 80  # 距离底部80像素
            
            click_positions = [
                (rect[0] + window_width // 2, base_y),  # 中央
                (rect[0] + window_width // 3, base_y),  # 左侧1/3
                (rect[0] + window_width * 2 // 3, base_y),  # 右侧2/3
            ]
            
            for pos_x, pos_y in click_positions:
                try:
                    self.progressed.emit(f"尝试点击输入区域: ({pos_x}, {pos_y})")
                    pyautogui.click(pos_x, pos_y)
                    time.sleep(0.3)
                    
                    # 验证点击是否成功：尝试输入一个字符然后删除
                    pyautogui.type('a')
                    time.sleep(0.1)
                    pyautogui.press('backspace')
                    time.sleep(0.1)
                    
                    self.progressed.emit(f"✅ 输入区域点击成功")
                    return True
                    
                except Exception as e:
                    self.progressed.emit(f"点击位置 ({pos_x}, {pos_y}) 失败: {e}")
                    continue
            
            return False
            
        except Exception as e:
            self.progressed.emit(f"智能点击输入区域失败: {e}")
            return False

    def _send_to_person(self, name: str, message: str, interval_min: float, interval_max: float):
        """发送消息给个人 - 增强版"""
        try:
            # 确保微信窗口激活
            if not self._activate_wechat():
                raise RuntimeError("无法激活微信窗口")
            
            # 搜索并进入聊天
            if not self._search_and_enter_chat(name):
                raise RuntimeError(f"无法找到或进入与{name}的聊天")
            
            # 查找输入框位置
            input_position = self._find_input_box_position()
            
            if input_position:
                # 使用找到的精确位置
                import pyautogui
                pyautogui.click(input_position[0], input_position[1])
                time.sleep(0.3)
            else:
                # 使用智能点击
                if not self._smart_click_input_area():
                    self.progressed.emit("⚠️ 无法定位输入框，使用默认位置")
            
            # 发送消息
            if self._send_message_content(message):
                self.progressed.emit(f"✅ 已发送给 {name}")
            else:
                raise RuntimeError("消息发送失败")
            
        except Exception as e:
            self.progressed.emit(f"❌ 发送给 {name} 失败: {e}")
            # 尝试备用方法
            self._send_via_hotkeys(name, message, interval_min, interval_max)

    def _send_message_content(self, message: str) -> bool:
        """发送消息内容"""
        try:
            if not self.HAS_PYAUTOGUI:
                return False
                
            import pyautogui
            import pyperclip
            
            # 复制消息到剪贴板并验证
            pyperclip.copy(message)
            time.sleep(0.2)
            
            # 验证剪贴板内容
            clipboard_content = pyperclip.paste()
            if clipboard_content != message:
                self.progressed.emit("⚠️ 剪贴板内容验证失败")
                return False
            
            # 清空输入框
            pyautogui.hotkey('ctrl', 'a')
            time.sleep(0.1)
            pyautogui.press('delete')
            time.sleep(0.1)
            
            # 粘贴消息
            pyautogui.hotkey('ctrl', 'v')
            time.sleep(0.5)
            
            # 发送消息
            pyautogui.press('enter')
            time.sleep(0.3)
            
            return True
            
        except Exception as e:
            self.progressed.emit(f"发送消息内容失败: {e}")
            return False

    def _send_via_hotkeys(self, name: str, message: str, interval_min: float, interval_max: float):
        """使用热键方式发送"""
        try:
            import pyautogui
            import pyperclip
            
            pyautogui.FAILSAFE = True
            pyautogui.PAUSE = 0.1
            
            # 确保微信窗口激活
            self._focus_wechat_window()
            time.sleep(1.0)
            
            # 搜索联系人
            pyautogui.hotkey('ctrl', 'f')
            time.sleep(0.5)
            pyperclip.copy(name)
            pyautogui.hotkey('ctrl', 'a')
            time.sleep(0.2)
            pyautogui.hotkey('ctrl', 'v')
            time.sleep(0.5)
            pyautogui.press('enter')
            time.sleep(1.5)
            
            # 发送消息
            pyperclip.copy(message)
            pyautogui.hotkey('ctrl', 'a')
            time.sleep(0.1)
            pyautogui.hotkey('ctrl', 'v')
            time.sleep(0.3)
            pyautogui.press('enter')
            
            self.progressed.emit(f"✅ 已发送给 {name} (热键方式)")
            
        except Exception as e:
            self.progressed.emit(f"❌ 发送给 {name} 完全失败: {e}")

    def _find_wechat_window(self):
        """查找微信窗口 - 改进版"""
        if not self.HAS_WIN32:
            return None
        
        import win32gui
        
        def enum_windows_callback(hwnd, windows):
            if not win32gui.IsWindowVisible(hwnd) or not win32gui.IsWindowEnabled(hwnd):
                return True
                
            window_text = win32gui.GetWindowText(hwnd)
            class_name = win32gui.GetClassName(hwnd)
            
            # 计算匹配得分
            match_score = 0
            
            # 主要匹配条件
            if class_name == "WeChatMainWndForPC":
                match_score += 50
            elif "WeChat" in class_name:
                match_score += 30
            elif "Wnd" in class_name and "PC" in class_name:
                match_score += 20
                
            # 窗口标题匹配
            if "微信" in window_text:
                match_score += 30
            elif "WeChat" in window_text:
                match_score += 25
                
            # Qt框架提示
            if "Qt5" in class_name or "Chrome_WidgetWin_1" in class_name:
                match_score += 10
                
            if match_score >= 30:  # 设定阈值
                windows.append((hwnd, match_score, window_text, class_name))
            
            return True
        
        windows = []
        try:
            win32gui.EnumWindows(enum_windows_callback, windows)
        except Exception as e:
            self.progressed.emit(f"枚举窗口失败: {e}")
            return None
        
        if not windows:
            return None
        
        # 按匹配得分排序，选择最佳匹配
        windows.sort(key=lambda x: x[1], reverse=True)
        best_match = windows[0]
        
        self.progressed.emit(f"找到微信窗口: {best_match[2]} (类名: {best_match[3]}, 得分: {best_match[1]})")
        return best_match[0]

    def _activate_wechat(self):
        """激活微信窗口 - 改进版"""
        try:
            if not self.HAS_WIN32:
                return False
                
            import win32gui, win32con
            
            # 重新查找微信窗口，确保窗口仍然有效
            self.wechat_hwnd = self._find_wechat_window()
            
            if not self.wechat_hwnd:
                self.progressed.emit("❌ 未找到微信窗口")
                return False
            
            # 检查窗口是否仍然有效
            try:
                if not win32gui.IsWindow(self.wechat_hwnd):
                    self.progressed.emit("❌ 微信窗口句柄无效")
                    return False
            except Exception:
                self.progressed.emit("❌ 无法验证微信窗口")
                return False
            
            # 多步骤激活窗口
            try:
                # 步骤1: 恢复窗口（如果最小化）
                win32gui.ShowWindow(self.wechat_hwnd, win32con.SW_RESTORE)
                time.sleep(0.2)
                
                # 步骤2: 设置为顶层窗口
                win32gui.SetWindowPos(self.wechat_hwnd, win32con.HWND_TOP, 0, 0, 0, 0, 
                                    win32con.SWP_NOMOVE | win32con.SWP_NOSIZE)
                time.sleep(0.2)
                
                # 步骤3: 设置为前台窗口
                win32gui.SetForegroundWindow(self.wechat_hwnd)
                time.sleep(0.3)
                
                # 验证激活是否成功
                try:
                    current_fg = win32gui.GetForegroundWindow()
                    if current_fg == self.wechat_hwnd:
                        self.progressed.emit("✅ 微信窗口已激活")
                        return True
                    else:
                        self.progressed.emit("⚠️ 微信窗口激活可能不完整")
                        return True  # 仍然尝试继续
                except Exception:
                    self.progressed.emit("⚠️ 无法验证窗口激活状态，继续尝试")
                    return True
                    
            except Exception as e:
                self.progressed.emit(f"❌ 激活微信窗口失败: {e}")
                return False
                
        except Exception as e:
            self.progressed.emit(f"❌ 激活微信窗口异常: {e}")
            return False

    def _focus_wechat_window(self):
        """激活微信窗口（兼容方法）"""
        return self._activate_wechat()

    def send_messages(self, messages: List[Dict], interval_min: float, interval_max: float, 
                     send_to_groups: bool = False, group_targets: List[str] = None,
                     test_mode: bool = False, test_target: str = "末"):
        """发送消息 - 支持个人和群聊"""
        try:
            if platform.system().lower() != "windows":
                raise RuntimeError("仅支持 Windows 平台")
            
            total_count = len(messages)
            
            if send_to_groups and group_targets:
                # 群聊模式：将所有消息汇总发送到指定群
                self._send_to_groups(messages, interval_min, interval_max, group_targets, test_mode, test_target)
            else:
                # 个人模式：逐一发送给个人
                self._send_to_individuals(messages, interval_min, interval_max, test_mode, test_target)
            
            self.finished.emit()
            
        except Exception as e:
            self.failed.emit(str(e))

    def _send_to_individuals(self, messages: List[Dict], interval_min: float, interval_max: float, 
                           test_mode: bool = False, test_target: str = "末"):
        """发送给个人"""
        total_count = len(messages)
        self.progressed.emit(f"开始个人发送，共 {total_count} 条消息")
        
        for i, msg_info in enumerate(messages):
            if self._stop.is_set():
                break
            
            original_name = msg_info['name']
            message = msg_info['message']
            
            # 测试模式处理
            if test_mode:
                target_name = test_target
                # 在消息前添加原始姓名信息
                test_message = f"[测试消息 - 原收件人: {original_name}]\n\n{message}"
            else:
                target_name = original_name
                test_message = message
            
            self.progressed.emit(f"正在发送 ({i+1}/{total_count}): {original_name}")
            
            try:
                self._send_to_person(target_name, test_message, interval_min, interval_max)
            except Exception as e:
                self.progressed.emit(f"❌ 发送失败: {e}")
            
            # 随机间隔，避免风控
            if i < len(messages) - 1:
                interval = random.uniform(interval_min, interval_max)
                self.progressed.emit(f"等待 {interval:.1f} 秒...")
                if not self._sleep(interval):
                    break

    def _send_to_groups(self, messages: List[Dict], interval_min: float, interval_max: float,
                       group_targets: List[str], test_mode: bool = False, test_target: str = "末"):
        """发送到群聊"""
        # 汇总所有消息为一条群消息
        summary_message = f"📊 今日用餐统计报告 ({datetime.now().strftime('%Y-%m-%d')})\n\n"
        summary_message += f"用餐人数：{len(messages)} 人\n\n"
        
        for i, msg_info in enumerate(messages, 1):
            name = msg_info['name']
            used_meals = msg_info['used_meals']
            remaining_meals = msg_info['remaining_meals']
            summary_message += f"{i}. {name}：已用{used_meals}次，剩余{remaining_meals}次\n"
        
        summary_message += f"\n💡 详细信息请查看餐数统计表"
        
        # 发送到每个选中的群
        targets = [test_target] if test_mode else group_targets
        
        for group_name in targets:
            if self._stop.is_set():
                break
                
            self.progressed.emit(f"正在发送到群: {group_name}")
            
            try:
                self._send_to_group(group_name, summary_message)
                self.progressed.emit(f"✅ 已发送到群: {group_name}")
            except Exception as e:
                self.progressed.emit(f"❌ 发送到群 {group_name} 失败: {e}")
            
            # 群间间隔
            if len(targets) > 1:
                interval = random.uniform(interval_min, interval_max)
                self.progressed.emit(f"等待 {interval:.1f} 秒...")
                if not self._sleep(interval):
                    break

    def _send_to_group(self, group_name: str, message: str):
        """发送消息到群聊"""
        try:
            # 确保微信窗口激活
            if not self._activate_wechat():
                raise RuntimeError("无法激活微信窗口")
            
            # 搜索并进入群聊
            if not self._search_and_enter_chat(group_name):
                raise RuntimeError(f"无法找到或进入群 {group_name}")
            
            # 查找输入框位置
            input_position = self._find_input_box_position()
            
            if input_position:
                # 使用找到的精确位置
                import pyautogui
                pyautogui.click(input_position[0], input_position[1])
                time.sleep(0.3)
            else:
                # 使用智能点击
                if not self._smart_click_input_area():
                    self.progressed.emit("⚠️ 无法定位输入框，使用默认位置")
            
            # 发送消息
            if not self._send_message_content(message):
                raise RuntimeError("消息发送失败")
                
        except Exception as e:
            raise RuntimeError(f"发送到群 {group_name} 失败: {e}")

    # 保持向后兼容
    def send_personal_messages(self, messages: List[Dict], interval_min: float, interval_max: float, test_mode: bool = False, test_target: str = "末"):
        """发送个人消息（向后兼容方法）"""
        return self.send_messages(messages, interval_min, interval_max, False, None, test_mode, test_target)


class MainWindow(QtWidgets.QMainWindow):
    """主窗口"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("简知轻食餐数统计发送器")
        self.setMinimumSize(1000, 720)
        
        self.df: Optional[pd.DataFrame] = None
        self.current_file: Optional[str] = None
        self.messages_to_send: List[Dict] = []
        
        self.sender = WeChatPersonalSender()
        self._send_thread: Optional[threading.Thread] = None
        
        self._init_ui()

    def _init_ui(self):
        """初始化UI"""
        central = QtWidgets.QWidget()
        self.setCentralWidget(central)
        root = QtWidgets.QVBoxLayout(central)
        root.setContentsMargins(16,16,16,16)
        root.setSpacing(12)

        # 文件选择区域
        drop = DropArea()
        drop.fileDropped.connect(self.on_file_dropped)
        root.addWidget(drop)

        row = QtWidgets.QHBoxLayout()
        self.file_label = QtWidgets.QLabel("未选择文件")
        pick = QtWidgets.QPushButton("选择扣餐表文件...")
        pick.clicked.connect(self.on_pick_file)
        row.addWidget(self.file_label, 1)
        row.addWidget(pick)
        root.addLayout(row)

        # 日期选择
        date_group = QtWidgets.QGroupBox("日期设置")
        date_layout = QtWidgets.QHBoxLayout(date_group)
        date_layout.addWidget(QtWidgets.QLabel("目标日期(1-31号):"))
        self.date_spin = QtWidgets.QSpinBox()
        self.date_spin.setRange(1, 31)
        self.date_spin.setValue(datetime.now().day)  # 默认今天
        date_layout.addWidget(self.date_spin)
        date_layout.addWidget(QtWidgets.QLabel("号"))
        date_layout.addStretch()
        root.addWidget(date_group)

        # 发送设置
        settings = QtWidgets.QGroupBox("发送设置")
        form = QtWidgets.QGridLayout(settings)
        
        # 发送间隔设置
        self.min_interval = QtWidgets.QDoubleSpinBox()
        self.min_interval.setRange(1.0, 10.0)
        self.min_interval.setSingleStep(0.1)
        self.min_interval.setValue(1.0)
        
        self.max_interval = QtWidgets.QDoubleSpinBox()
        self.max_interval.setRange(1.0, 10.0)
        self.max_interval.setSingleStep(0.1)
        self.max_interval.setValue(2.0)
        
        form.addWidget(QtWidgets.QLabel("发送间隔（秒）:"), 0, 0)
        h = QtWidgets.QHBoxLayout()
        h.addWidget(QtWidgets.QLabel("最小"))
        h.addWidget(self.min_interval)
        h.addSpacing(8)
        h.addWidget(QtWidgets.QLabel("最大"))
        h.addWidget(self.max_interval)
        w = QtWidgets.QWidget()
        w.setLayout(h)
        form.addWidget(w, 0, 1, 1, 3)
        
        # 发送方式选择
        send_method_group = QtWidgets.QGroupBox("发送方式")
        send_method_layout = QtWidgets.QHBoxLayout(send_method_group)
        
        self.send_individual = QtWidgets.QRadioButton("发送给个人")
        self.send_to_groups = QtWidgets.QRadioButton("发送到群聊")
        self.send_individual.setChecked(True)  # 默认发送给个人
        
        send_method_layout.addWidget(self.send_individual)
        send_method_layout.addWidget(self.send_to_groups)
        send_method_layout.addStretch()
        
        form.addWidget(send_method_group, 1, 0, 1, 4)
        
        # 群聊设置（当选择发送到群聊时启用）
        group_settings = QtWidgets.QGroupBox("群聊设置")
        group_layout = QtWidgets.QGridLayout(group_settings)
        
        self.group1_name = QtWidgets.QLineEdit("简知用餐群1")
        self.group2_name = QtWidgets.QLineEdit("简知用餐群2")
        self.send_to_group1 = QtWidgets.QCheckBox("发送到群1")
        self.send_to_group2 = QtWidgets.QCheckBox("发送到群2")
        self.send_to_group1.setChecked(True)
        self.send_to_group2.setChecked(True)
        
        group_layout.addWidget(QtWidgets.QLabel("群1名称:"), 0, 0)
        group_layout.addWidget(self.group1_name, 0, 1)
        group_layout.addWidget(self.send_to_group1, 0, 2)
        group_layout.addWidget(QtWidgets.QLabel("群2名称:"), 1, 0)
        group_layout.addWidget(self.group2_name, 1, 1)
        group_layout.addWidget(self.send_to_group2, 1, 2)
        
        form.addWidget(group_settings, 2, 0, 1, 4)
        
        # 测试模式
        test_group = QtWidgets.QGroupBox("测试模式")
        test_layout = QtWidgets.QHBoxLayout(test_group)
        
        self.test_mode = QtWidgets.QCheckBox("测试模式")
        self.test_target = QtWidgets.QLineEdit("末")
        
        test_layout.addWidget(self.test_mode)
        test_layout.addWidget(QtWidgets.QLabel("测试发送至:"))
        test_layout.addWidget(self.test_target)
        test_layout.addStretch()
        
        form.addWidget(test_group, 3, 0, 1, 4)
        
        # 连接信号，控制群聊设置的启用状态
        self.send_individual.toggled.connect(self._on_send_method_changed)
        self.send_to_groups.toggled.connect(self._on_send_method_changed)
        
        # 初始状态：禁用群聊设置
        group_settings.setEnabled(False)
        self.group_settings_widget = group_settings
        
        root.addWidget(settings)

        # 操作按钮
        actions = QtWidgets.QHBoxLayout()
        self.btn_analyze = QtWidgets.QPushButton("分析数据")
        self.btn_analyze.clicked.connect(self.on_analyze)
        self.btn_send = QtWidgets.QPushButton("开始发送")
        self.btn_send.clicked.connect(self.on_send)
        self.btn_stop = QtWidgets.QPushButton("停止发送")
        self.btn_stop.clicked.connect(self.on_stop)
        
        actions.addWidget(self.btn_analyze)
        actions.addWidget(self.btn_send)
        actions.addWidget(self.btn_stop)
        actions.addStretch(1)
        root.addLayout(actions)
        
        # 测试功能按钮
        test_actions = QtWidgets.QHBoxLayout()
        self.btn_test_search = QtWidgets.QPushButton("测试联系人搜索")
        self.btn_test_send = QtWidgets.QPushButton("测试发送消息")
        self.btn_test_input = QtWidgets.QPushButton("测试输入框定位")
        
        self.btn_test_search.clicked.connect(self.on_test_search)
        self.btn_test_send.clicked.connect(self.on_test_send)
        self.btn_test_input.clicked.connect(self.on_test_input)
        
        test_actions.addWidget(QtWidgets.QLabel("测试功能:"))
        test_actions.addWidget(self.btn_test_search)
        test_actions.addWidget(self.btn_test_send)
        test_actions.addWidget(self.btn_test_input)
        test_actions.addStretch(1)
        root.addLayout(test_actions)

        # 预览和日志区域
        self.preview = QtWidgets.QPlainTextEdit()
        self.preview.setReadOnly(True)
        root.addWidget(self.preview, 1)
        
        self.status = QtWidgets.QLabel("")
        self.status.setStyleSheet("color:#0a7;")
        root.addWidget(self.status)

        # 初始状态
        self.btn_send.setEnabled(False)
        
        # 全局热键
        self._setup_hotkey()

    def _setup_hotkey(self):
        """设置全局热键"""
        t = threading.Thread(target=self._hotkey_worker, daemon=True)
        t.start()

    def _hotkey_worker(self):
        """热键监听线程"""
        try:
            import keyboard
            keyboard.add_hotkey("ctrl+shift+s", lambda: self.on_stop())
            while True:
                time.sleep(1)
        except Exception:
            return

    def on_pick_file(self):
        """选择文件"""
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "选择扣餐表Excel文件", 
            os.path.expanduser("~"), 
            "Excel文件 (*.xlsx *.xls *.xlsm)"
        )
        if path:
            self._load_file(path)

    def on_file_dropped(self, path: str):
        """文件拖拽处理"""
        self._load_file(path)

    def _load_file(self, path: str):
        """加载文件"""
        try:
            df, sheets = load_excel_file(path)
            df = normalize_columns(df)
            self.df = df
            self.current_file = path
            self.file_label.setText(f"已加载：{os.path.basename(path)}")
            self.status.setText("文件加载成功，请点击'分析数据'")
            
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "加载失败", f"{e}\n\n{traceback.format_exc()}")

    def on_analyze(self):
        """分析数据"""
        try:
            if self.df is None:
                raise RuntimeError("请先加载扣餐表文件")
            
            target_date = self.date_spin.value()
            messages, summary = analyze_meal_data(self.df, target_date)
            
            self.messages_to_send = messages
            self.preview.setPlainText(summary)
            
            if messages:
                self.btn_send.setEnabled(True)
                self.status.setText(f"分析完成，找到 {len(messages)} 位今日用餐用户")
            else:
                self.btn_send.setEnabled(False)
                self.status.setText(f"{target_date}号暂无用餐记录")
                
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "分析失败", str(e))

    def on_send(self):
        """开始发送"""
        try:
            if not self.messages_to_send:
                raise RuntimeError("没有可发送的消息，请先分析数据")
            
            mi = float(self.min_interval.value())
            ma = float(self.max_interval.value())
            if ma < mi:
                QtWidgets.QMessageBox.warning(self, "参数错误", "最大间隔不能小于最小间隔")
                return
            
            test_mode = self.test_mode.isChecked()
            test_target = self.test_target.text().strip()
            
            if test_mode and not test_target:
                QtWidgets.QMessageBox.warning(self, "参数错误", "测试模式需要指定测试目标")
                return
            
            # 检查发送方式
            send_to_groups = self.send_to_groups.isChecked()
            group_targets = []
            
            if send_to_groups:
                # 群聊模式，检查选中的群
                if self.send_to_group1.isChecked():
                    group1_name = self.group1_name.text().strip()
                    if group1_name:
                        group_targets.append(group1_name)
                
                if self.send_to_group2.isChecked():
                    group2_name = self.group2_name.text().strip()
                    if group2_name:
                        group_targets.append(group2_name)
                
                if not group_targets:
                    QtWidgets.QMessageBox.warning(self, "参数错误", "群聊模式需要选择至少一个群")
                    return
            
            # 确认发送
            if send_to_groups:
                if test_mode:
                    confirm_msg = f"即将以群聊模式发送统计报告到测试群 '{test_target}'"
                else:
                    target_groups = ", ".join(group_targets)
                    confirm_msg = f"即将以群聊模式发送统计报告到: {target_groups}"
                confirm_msg += f"\n\n将汇总 {len(self.messages_to_send)} 人的用餐数据为一条群消息"
            else:
                if test_mode:
                    confirm_msg = f"即将以个人模式发送 {len(self.messages_to_send)} 条消息到测试目标 '{test_target}'"
                else:
                    confirm_msg = f"即将以个人模式逐一发送 {len(self.messages_to_send)} 条个人消息"
            
            confirm_msg += f"\n\n发送间隔: {mi}-{ma} 秒\n\n确认开始发送吗？"
            
            result = QtWidgets.QMessageBox.question(
                self, "确认发送", confirm_msg,
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.No
            )
            
            if result != QtWidgets.QMessageBox.Yes:
                return
            
            self.btn_send.setEnabled(False)
            self.sender.progressed.connect(self._on_progress)
            self.sender.finished.connect(self._on_finished)
            self.sender.failed.connect(self._on_failed)
            self.sender._stop.clear()
            
            self._send_thread = threading.Thread(
                target=self.sender.send_messages,
                args=(self.messages_to_send, mi, ma, send_to_groups, group_targets, test_mode, test_target),
                daemon=True
            )
            self._send_thread.start()
            
            if send_to_groups:
                self.status.setText("正在发送群聊统计...")
            else:
                self.status.setText("正在发送个人消息...")
            
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "发送失败", str(e))

    def on_stop(self):
        """停止发送"""
        try:
            self.sender.stop()
            self.status.setText("停止指令已发送")
        except Exception:
            pass

    def _on_progress(self, msg: str):
        """发送进度更新"""
        self.status.setText(msg)

    def _on_finished(self):
        """发送完成"""
        self.btn_send.setEnabled(True)
        self.status.setText("发送完成！")

    def _on_failed(self, err: str):
        """发送失败"""
        self.btn_send.setEnabled(True)
        QtWidgets.QMessageBox.critical(self, "发送失败", err)

    def _on_send_method_changed(self):
        """发送方式改变时的处理"""
        if self.send_to_groups.isChecked():
            self.group_settings_widget.setEnabled(True)
        else:
            self.group_settings_widget.setEnabled(False)

    def on_test_search(self):
        """测试联系人搜索"""
        try:
            test_target = self.test_target.text().strip()
            if not test_target:
                QtWidgets.QMessageBox.warning(self, "警告", "请输入测试目标联系人")
                return
            
            self.status.setText("正在测试联系人搜索...")
            
            # 在新线程中测试
            def test_search():
                try:
                    if not self.sender._activate_wechat():
                        self.sender.progressed.emit("❌ 无法激活微信窗口")
                        return
                    
                    success = self.sender._search_and_enter_chat(test_target)
                    if success:
                        self.sender.progressed.emit(f"✅ 成功找到并进入与 {test_target} 的聊天")
                    else:
                        self.sender.progressed.emit(f"❌ 无法找到联系人 {test_target}")
                        
                except Exception as e:
                    self.sender.progressed.emit(f"❌ 测试搜索失败: {e}")
            
            self.sender.progressed.connect(self._on_progress)
            thread = threading.Thread(target=test_search, daemon=True)
            thread.start()
            
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "测试失败", str(e))

    def on_test_send(self):
        """测试发送消息"""
        try:
            test_target = self.test_target.text().strip()
            if not test_target:
                QtWidgets.QMessageBox.warning(self, "警告", "请输入测试目标联系人")
                return
            
            test_message = f"[测试消息] 这是来自餐数统计发送器的测试消息，发送时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            self.status.setText("正在测试发送消息...")
            
            # 在新线程中测试
            def test_send():
                try:
                    self.sender._send_to_person(test_target, test_message, 1.0, 2.0)
                except Exception as e:
                    self.sender.progressed.emit(f"❌ 测试发送失败: {e}")
            
            self.sender.progressed.connect(self._on_progress)
            thread = threading.Thread(target=test_send, daemon=True)
            thread.start()
            
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "测试失败", str(e))

    def on_test_input(self):
        """测试输入框定位"""
        try:
            self.status.setText("正在测试输入框定位...")
            
            # 在新线程中测试
            def test_input():
                try:
                    if not self.sender._activate_wechat():
                        self.sender.progressed.emit("❌ 无法激活微信窗口")
                        return
                    
                    # 测试各种定位方法
                    self.sender.progressed.emit("🔍 测试控件识别方法...")
                    pos1 = self.sender._find_input_by_control()
                    if pos1:
                        self.sender.progressed.emit(f"✅ 控件识别找到位置: {pos1}")
                        self._mark_position(pos1, "控件识别", "red")
                    else:
                        self.sender.progressed.emit("❌ 控件识别未找到输入框")
                    
                    time.sleep(1)
                    
                    self.sender.progressed.emit("🔍 测试窗口计算方法...")
                    pos2 = self.sender._find_input_by_window_calc()
                    if pos2:
                        self.sender.progressed.emit(f"✅ 窗口计算找到位置: {pos2}")
                        self._mark_position(pos2, "窗口计算", "blue")
                    else:
                        self.sender.progressed.emit("❌ 窗口计算未找到输入框")
                    
                    time.sleep(1)
                    
                    self.sender.progressed.emit("🔍 测试智能点击方法...")
                    success = self.sender._smart_click_input_area()
                    if success:
                        self.sender.progressed.emit("✅ 智能点击成功")
                    else:
                        self.sender.progressed.emit("❌ 智能点击失败")
                        
                except Exception as e:
                    self.sender.progressed.emit(f"❌ 测试输入框定位失败: {e}")
            
            self.sender.progressed.connect(self._on_progress)
            thread = threading.Thread(target=test_input, daemon=True)
            thread.start()
            
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "测试失败", str(e))

    def _mark_position(self, position, method_name, color):
        """在屏幕上标记位置"""
        try:
            import tkinter as tk
            
            def show_marker():
                marker = tk.Toplevel()
                marker.title(f"{method_name}检测位置")
                marker.geometry(f"8x8+{position[0]-4}+{position[1]-4}")
                marker.configure(bg=color)
                marker.attributes('-topmost', True)
                marker.overrideredirect(True)
                
                # 3秒后自动关闭
                marker.after(3000, marker.destroy)
                marker.mainloop()
            
            # 在新线程中显示标记，避免阻塞主线程
            marker_thread = threading.Thread(target=show_marker, daemon=True)
            marker_thread.start()
            
        except Exception as e:
            self.sender.progressed.emit(f"标记位置失败: {e}")


def main():
    """主函数"""
    # Windows编码设置
    if platform.system().lower() == "windows":
        try:
            import ctypes
            # 设置控制台编码为UTF-8
            ctypes.windll.kernel32.SetConsoleOutputCP(65001)
            ctypes.windll.kernel32.SetConsoleCP(65001)
        except Exception:
            pass
        
        # 设置环境变量
        os.environ['PYTHONIOENCODING'] = 'utf-8'
        
        # 设置locale
        try:
            import locale
            locale.setlocale(locale.LC_ALL, 'Chinese (Simplified)_China.utf8')
        except Exception:
            try:
                locale.setlocale(locale.LC_ALL, 'zh_CN.UTF-8')
            except Exception:
                pass
    
    # 设置Qt应用程序属性
    app = QtWidgets.QApplication(sys.argv)
    
    # 设置应用程序编码
    try:
        app.setStyle('Fusion')  # 使用Fusion样式，更好的中文支持
    except Exception:
        pass
    
    # 设置字体，确保中文显示正常
    try:
        font = QtGui.QFont()
        font.setFamily("Microsoft YaHei")  # 微软雅黑
        font.setPointSize(9)
        app.setFont(font)
    except Exception:
        pass
    
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
