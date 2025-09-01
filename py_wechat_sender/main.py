import os
import sys
import time
import random
import threading
import platform
import traceback
from typing import List, Optional, Tuple, Dict

import pandas as pd
from PyQt5 import QtCore, QtGui, QtWidgets


REQUIRED_COLUMNS = ["商品信息", "支付状态", "订单状态", "收货地址", "用户备注"]


def detect_csv_encoding(file_path: str) -> str:
    try:
        import chardet  # type: ignore
    except Exception:
        return "utf-8"
    with open(file_path, "rb") as f:
        raw = f.read(4096)
    result = chardet.detect(raw)
    return result.get("encoding") or "utf-8"


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    def _to_str(col) -> str:
        if isinstance(col, tuple):
            return " ".join([str(x).strip() for x in col])
        return str(col).strip()
    df.columns = [_to_str(c) for c in df.columns]
    return df


def load_dataframe(file_path: str) -> Tuple[pd.DataFrame, List[str]]:
    ext = os.path.splitext(file_path)[1].lower()
    if ext in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        from openpyxl import load_workbook
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        # header guess
        header_idx = 0
        best = -1
        for i in range(min(10, len(rows))):
            cnt = sum(1 for v in rows[i] if v not in (None, ""))
            if cnt > best:
                best = cnt
                header_idx = i
        headers = [str(h).strip() if h is not None else f"列{i+1}" for i, h in enumerate(rows[header_idx] if rows else [])]
        data = rows[header_idx+1:]
        width = len(headers)
        norm = []
        for r in data:
            r = list(r)
            if len(r) < width:
                r += [None]*(width-len(r))
            elif len(r) > width:
                r = r[:width]
            norm.append(r)
        return pd.DataFrame(norm, columns=headers), sheets
    if ext == ".xls":
        from pyexcel_xls import get_data  # type: ignore
        data = get_data(file_path)
        sheets = list(data.keys())
        rows = data[sheets[0]]
        header_idx = 0
        best = -1
        for i in range(min(10, len(rows))):
            cnt = sum(1 for v in rows[i] if v not in (None, ""))
            if cnt > best:
                best = cnt
                header_idx = i
        headers = [str(h).strip() if h is not None else f"列{i+1}" for i, h in enumerate(rows[header_idx] if rows else [])]
        data_rows = rows[header_idx+1:]
        width = len(headers)
        norm = []
        for r in data_rows:
            r = list(r)
            if len(r) < width:
                r += [None]*(width-len(r))
            elif len(r) > width:
                r = r[:width]
            norm.append(r)
        return pd.DataFrame(norm, columns=headers), sheets
    if ext == ".xlsb":
        xls = pd.ExcelFile(file_path, engine="pyxlsb")
        sheets = xls.sheet_names
        df = pd.read_excel(xls, sheet_name=sheets[0])
        return df, sheets
    if ext == ".ods":
        xls = pd.ExcelFile(file_path, engine="odf")
        sheets = xls.sheet_names
        df = pd.read_excel(xls, sheet_name=sheets[0])
        return df, sheets
    if ext in [".csv", ".txt"]:
        enc = detect_csv_encoding(file_path)
        try:
            df = pd.read_csv(file_path, encoding=enc, sep=None, engine="python")
        except Exception:
            df = pd.read_csv(file_path, encoding="utf-8", sep=",", engine="python", errors="ignore")
        return df, ["CSV"]
    # last resort try parse as csv
    enc = detect_csv_encoding(file_path)
    df = pd.read_csv(file_path, encoding=enc, sep=None, engine="python")
    return df, ["CSV"]


def infer_default_mapping(df: pd.DataFrame) -> Dict[str, str]:
    cols = list(df.columns)
    mapping: Dict[str, str] = {}
    for target in REQUIRED_COLUMNS:
        if target in cols:
            mapping[target] = target
        else:
            k = target.replace(" ", "")
            found = None
            for c in cols:
                if str(c).replace(" ", "") == k:
                    found = c
                    break
            mapping[target] = found or cols[0]
    return mapping


def split_address(text: str) -> str:
    if not isinstance(text, str):
        text = str(text) if pd.notna(text) else ""
    t = text.strip()
    if not t:
        return " -  - "
    parts = [p.strip() for p in re_split_hyphen(t)]
    if len(parts) >= 3:
        return f"{parts[0]} - {parts[1]} - {' - '.join(parts[2:])}"
    m = find_phone_like(t)
    if m:
        s, e = m.span()
        name = t[:s].strip(" -")
        phone = t[s:e]
        addr = t[e:].strip(" -")
        return f"{name} - {phone} - {addr}" if (name or addr) else f" - {phone} - "
    return f" -  - {t}"


def re_split_hyphen(text: str) -> List[str]:
    import re
    return re.split(r"\s*[-\-\–\—\－]\s*", text)


def find_phone_like(text: str):
    import re
    return re.search(r"(\d[\d\s-]{5,19}\d)", text)


def filter_and_order(df: pd.DataFrame, mapping: Dict[str, str]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    df = df.copy()
    df["__row__"] = range(1, len(df) + 1)
    for col in mapping.values():
        if col in df.columns:
            df[col] = df[col].fillna("")
    paid = df[df[mapping["支付状态"]].astype(str).str.strip() == "已支付"]
    def excluded(row):
        st = str(row[mapping["订单状态"]]).strip()
        pay = str(row[mapping["支付状态"]]).strip()
        if pay == "未支付":
            return True
        if st in ("已取消", "用户申请退款"):
            return True
        if pay == "已退款":
            return True
        return False
    eff = paid[~paid.apply(excluded, axis=1)].copy()
    lunch = eff[eff[mapping["商品信息"]].astype(str).str.strip() == "明日午餐 x1"].copy()
    dinner = eff[eff[mapping["商品信息"]].astype(str).str.strip() == "明日晚餐 x1"].copy()
    lunch = lunch.sort_values("__row__", ascending=False)
    dinner = dinner.sort_values("__row__", ascending=False)
    return lunch, dinner


def build_output(df: pd.DataFrame, mapping: Dict[str, str], start: int, title: str, product_label: str) -> str:
    lines: List[str] = []
    lines.append(f"### {title}（商品信息：{product_label}，编号从{start}开始）")
    cur = start
    for _, row in df.iterrows():
        addr = split_address(str(row.get(mapping["收货地址"], "")))
        note = str(row.get(mapping["用户备注"], ""))
        lines.append(str(cur))
        lines.append(addr)
        if note.strip():
            lines.append(f"（用户备注：{note}）")
        cur += 1
    return "\n".join(lines)


class DropArea(QtWidgets.QFrame):
    fileDropped = QtCore.pyqtSignal(str)
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.setStyleSheet("""
            QFrame { border: 2px dashed #bbb; border-radius: 8px; background: #fafafa; }
        """)
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(16,16,16,16)
        label = QtWidgets.QLabel("将 Excel/CSV 文件拖拽到此处，或点击下方按钮选择文件")
        label.setAlignment(QtCore.Qt.AlignCenter)
        label.setStyleSheet("color:#555;font-size:14px;")
        layout.addWidget(label)
    def dragEnterEvent(self, e: QtGui.QDragEnterEvent) -> None:
        e.acceptProposedAction() if e.mimeData().hasUrls() else e.ignore()
    def dropEvent(self, e: QtGui.QDropEvent) -> None:
        urls = e.mimeData().urls()
        if urls:
            self.fileDropped.emit(urls[0].toLocalFile())


class WeChatSender(QtCore.QObject):
    progressed = QtCore.pyqtSignal(str)
    finished = QtCore.pyqtSignal()
    failed = QtCore.pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self._stop = threading.Event()

    def stop(self):
        self._stop.set()

    def _sleep(self, seconds: float) -> bool:
        end = time.time() + seconds
        while time.time() < end:
            if self._stop.is_set():
                return False
            time.sleep(0.05)
        return True

    def _ensure_wechat(self):
        # Try attach to WeChat window; if not, try to start
        import uiautomation as auto
        main = auto.WindowControl(searchDepth=1, ClassName="WeChatMainWndForPC")
        if not main.Exists(0.5):
            try:
                os.startfile("WeChat.exe")
            except Exception:
                pass
            auto.WaitForExistence(main, 15.0)
        if not main.Exists(0.5):
            raise RuntimeError("未找到微信窗口，请先登录微信")
        return main

    def _send_to_group(self, group: str, text: str, interval_min: float, interval_max: float):
        import uiautomation as auto
        import pyperclip
        main = self._ensure_wechat()
        main.SetActive()
        # Focus search
        # In recent versions, there is a search edit control with name like 搜索
        search = main.EditControl(foundIndex=1)
        if not search.Exists(0.5):
            raise RuntimeError("未找到微信搜索框")
        search.Click()
        search.SendKeys("^a{Delete}")
        search.SendKeys(group)
        time.sleep(0.5)
        search.SendKeys("{Enter}")
        time.sleep(0.5)
        # Chat input is a RICHEDIT control, usually last EditControl
        edits = main.EditControls()
        if not edits:
            raise RuntimeError("未找到输入框")
        input_box = edits[-1]
        # Paste chunks and send
        chunks = split_message_chunks(text, 3500)
        for idx, chunk in enumerate(chunks, start=1):
            if self._stop.is_set():
                return
            pyperclip.copy(chunk)
            input_box.Click()
            input_box.SendKeys("^v")
            input_box.SendKeys("{Enter}")
            self.progressed.emit(f"已发送 {group} 第 {idx} 段")
            if idx < len(chunks):
                d = random.uniform(interval_min, interval_max)
                if not self._sleep(d):
                    return

    def send(self, items: List[Tuple[str, str]], interval_min: float, interval_max: float):
        try:
            if platform.system().lower() != "windows":
                raise RuntimeError("仅支持 Windows 平台")
            for i, (group, text) in enumerate(items):
                if self._stop.is_set():
                    break
                group = str(group).strip()
                if not group:
                    continue
                self.progressed.emit(f"正在发送到：{group}")
                try:
                    self._send_to_group(group, text, interval_min, interval_max)
                except Exception as e:
                    self.progressed.emit(f"发送到 {group} 失败：{e}")
                if i < len(items) - 1:
                    d = random.uniform(interval_min, interval_max)
                    if not self._sleep(d):
                        break
            self.finished.emit()
        except Exception as e:
            self.failed.emit(str(e))


def split_message_chunks(text: str, max_len: int = 3500) -> List[str]:
    if len(text) <= max_len:
        return [text]
    parts: List[str] = []
    current: List[str] = []
    current_len = 0
    for line in text.splitlines():
        add_len = len(line) + (1 if current else 0)
        if current_len + add_len > max_len:
            parts.append("\n".join(current))
            current = [line]
            current_len = len(line)
        else:
            if current:
                current.append(line)
                current_len += len(line) + 1
            else:
                current = [line]
                current_len = len(line)
    if current:
        parts.append("\n".join(current))
    return parts


class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("简知订单整理发送器（Windows自动化）")
        self.setMinimumSize(960, 680)
        self.df: Optional[pd.DataFrame] = None
        self.current_file: Optional[str] = None
        self.mapping: Optional[Dict[str, str]] = None

        self.sender = WeChatSender()
        self._send_thread: Optional[threading.Thread] = None

        self._init_ui()

    def _init_ui(self):
        central = QtWidgets.QWidget()
        self.setCentralWidget(central)
        root = QtWidgets.QVBoxLayout(central)
        root.setContentsMargins(16,16,16,16)
        root.setSpacing(12)

        drop = DropArea()
        drop.fileDropped.connect(self.on_file_dropped)
        root.addWidget(drop)

        row = QtWidgets.QHBoxLayout()
        self.file_label = QtWidgets.QLabel("未选择文件")
        pick = QtWidgets.QPushButton("选择文件…")
        pick.clicked.connect(self.on_pick_file)
        row.addWidget(self.file_label, 1)
        row.addWidget(pick)
        root.addLayout(row)

        map_group = QtWidgets.QGroupBox("字段映射（自动识别，可手动调整）")
        grid = QtWidgets.QGridLayout(map_group)
        self.cmb_product = QtWidgets.QComboBox()
        self.cmb_pay = QtWidgets.QComboBox()
        self.cmb_status = QtWidgets.QComboBox()
        self.cmb_addr = QtWidgets.QComboBox()
        self.cmb_note = QtWidgets.QComboBox()
        for i, (label, w) in enumerate([
            ("商品信息", self.cmb_product),
            ("支付状态", self.cmb_pay),
            ("订单状态", self.cmb_status),
            ("收货地址", self.cmb_addr),
            ("用户备注", self.cmb_note),
        ]):
            grid.addWidget(QtWidgets.QLabel(label+"："), i, 0)
            grid.addWidget(w, i, 1)
        map_group.setEnabled(False)
        self.map_group = map_group
        root.addWidget(map_group)

        settings = QtWidgets.QGroupBox("发送设置")
        form = QtWidgets.QGridLayout(settings)
        self.lunch_start = QtWidgets.QSpinBox(); self.lunch_start.setRange(1, 100000); self.lunch_start.setValue(7)
        self.dinner_start = QtWidgets.QSpinBox(); self.dinner_start.setRange(1, 100000); self.dinner_start.setValue(7)
        self.cmb_lunch_group = QtWidgets.QComboBox(); self.cmb_lunch_group.setEditable(True); self.cmb_lunch_group.addItems(["简知午餐群", "末"])
        self.cmb_dinner_group = QtWidgets.QComboBox(); self.cmb_dinner_group.setEditable(True); self.cmb_dinner_group.addItems(["简知晚餐群", "末"])
        self.min_interval = QtWidgets.QDoubleSpinBox(); self.min_interval.setRange(0.1, 10.0); self.min_interval.setSingleStep(0.1); self.min_interval.setValue(1.0)
        self.max_interval = QtWidgets.QDoubleSpinBox(); self.max_interval.setRange(0.1, 10.0); self.max_interval.setSingleStep(0.1); self.max_interval.setValue(1.5)
        self.test_mode = QtWidgets.QCheckBox("测试模式（发送到：末）")

        form.addWidget(QtWidgets.QLabel("午餐起始编号："), 0, 0); form.addWidget(self.lunch_start, 0, 1)
        form.addWidget(QtWidgets.QLabel("晚餐起始编号："), 0, 2); form.addWidget(self.dinner_start, 0, 3)
        form.addWidget(QtWidgets.QLabel("午餐发送至："), 1, 0); form.addWidget(self.cmb_lunch_group, 1, 1)
        form.addWidget(QtWidgets.QLabel("晚餐发送至："), 1, 2); form.addWidget(self.cmb_dinner_group, 1, 3)
        form.addWidget(QtWidgets.QLabel("发送间隔（秒）："), 2, 0)
        h = QtWidgets.QHBoxLayout(); h.addWidget(QtWidgets.QLabel("最小")); h.addWidget(self.min_interval); h.addSpacing(8); h.addWidget(QtWidgets.QLabel("最大")); h.addWidget(self.max_interval)
        w = QtWidgets.QWidget(); w.setLayout(h); form.addWidget(w, 2, 1, 1, 3)
        form.addWidget(self.test_mode, 3, 0, 1, 4)
        root.addWidget(settings)

        actions = QtWidgets.QHBoxLayout()
        self.btn_preview = QtWidgets.QPushButton("预览"); self.btn_preview.clicked.connect(self.on_preview)
        self.btn_send = QtWidgets.QPushButton("发送"); self.btn_send.clicked.connect(self.on_send)
        self.btn_stop = QtWidgets.QPushButton("停止（Ctrl+Shift+S）"); self.btn_stop.clicked.connect(self.on_stop)
        actions.addWidget(self.btn_preview); actions.addWidget(self.btn_send); actions.addWidget(self.btn_stop); actions.addStretch(1)
        root.addLayout(actions)

        self.preview = QtWidgets.QPlainTextEdit(); self.preview.setReadOnly(True)
        root.addWidget(self.preview, 1)
        self.status = QtWidgets.QLabel(""); self.status.setStyleSheet("color:#0a7;")
        root.addWidget(self.status)

        # Global hotkey
        t = threading.Thread(target=self._hotkey_worker, daemon=True)
        t.start()

    def _hotkey_worker(self):
        try:
            import keyboard  # type: ignore
            keyboard.add_hotkey("ctrl+shift+s", lambda: self.on_stop())
            while True:
                time.sleep(1)
        except Exception:
            return

    def on_pick_file(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "选择 Excel/CSV 文件", os.path.expanduser("~"), "表格文件 (*.xlsx *.xls *.xlsb *.ods *.csv *.txt)")
        if path:
            self._load_file(path)

    def on_file_dropped(self, path: str):
        self._load_file(path)

    def _load_file(self, path: str):
        try:
            df, _ = load_dataframe(path)
            df = normalize_columns(df)
            self.df = df
            self.current_file = path
            self.file_label.setText(f"已加载：{os.path.basename(path)}")
            self.map_group.setEnabled(True)
            for cmb in [self.cmb_product, self.cmb_pay, self.cmb_status, self.cmb_addr, self.cmb_note]:
                cmb.clear(); cmb.addItems([str(c) for c in self.df.columns])
            m = infer_default_mapping(self.df)
            self.cmb_product.setCurrentText(m["商品信息"])
            self.cmb_pay.setCurrentText(m["支付状态"])
            self.cmb_status.setCurrentText(m["订单状态"])
            self.cmb_addr.setCurrentText(m["收货地址"])
            self.cmb_note.setCurrentText(m["用户备注"])
            self.status.setText("文件加载成功。")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "加载失败", f"{e}\n\n{traceback.format_exc()}")

    def _mapping(self) -> Dict[str, str]:
        return {
            "商品信息": self.cmb_product.currentText(),
            "支付状态": self.cmb_pay.currentText(),
            "订单状态": self.cmb_status.currentText(),
            "收货地址": self.cmb_addr.currentText(),
            "用户备注": self.cmb_note.currentText(),
        }

    def _build_texts(self) -> Tuple[str, str]:
        if self.df is None:
            raise RuntimeError("请先加载 Excel/CSV 文件")
        mp = self._mapping()
        lunch, dinner = filter_and_order(self.df, mp)
        lunch_text = build_output(lunch, mp, self.lunch_start.value(), "一、午餐", "明日午餐 x1")
        dinner_text = build_output(dinner, mp, self.dinner_start.value(), "二、晚餐", "明日晚餐 x1")
        return lunch_text, dinner_text

    def on_preview(self):
        try:
            lunch_text, dinner_text = self._build_texts()
            self.preview.setPlainText((lunch_text + "\n\n" + dinner_text).strip())
            self.status.setText("预览已生成。")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "预览失败", str(e))

    def on_send(self):
        try:
            lunch_text, dinner_text = self._build_texts()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "发送失败", str(e))
            return
        mi = float(self.min_interval.value()); ma = float(self.max_interval.value())
        if ma < mi:
            QtWidgets.QMessageBox.warning(self, "参数错误", "最大发送间隔不能小于最小发送间隔")
            return
        lunch_group = self.cmb_lunch_group.currentText().strip()
        dinner_group = self.cmb_dinner_group.currentText().strip()
        items: List[Tuple[str, str]] = []
        test = self.test_mode.isChecked()
        items.append((("末" if test else lunch_group), lunch_text))
        items.append((("末" if test else dinner_group), dinner_text))
        items = [(g, t) for g, t in items if g]
        if not items:
            QtWidgets.QMessageBox.warning(self, "缺少群聊", "请至少设置一个群聊或开启测试模式")
            return
        self.btn_send.setEnabled(False)
        self.sender.progressed.connect(self._on_progress)
        self.sender.finished.connect(self._on_finished)
        self.sender.failed.connect(self._on_failed)
        self.sender._stop.clear()
        self._send_thread = threading.Thread(target=self.sender.send, args=(items, mi, ma), daemon=True)
        self._send_thread.start()
        self.status.setText("正在发送…")

    def on_stop(self):
        try:
            self.sender.stop()
            self.status.setText("停止指令已发送，将尽快停止。")
        except Exception:
            pass

    def _on_progress(self, msg: str):
        self.status.setText(msg)

    def _on_finished(self):
        self.btn_send.setEnabled(True)
        self.status.setText("发送完成。")

    def _on_failed(self, err: str):
        self.btn_send.setEnabled(True)
        QtWidgets.QMessageBox.critical(self, "发送失败", err)


def main():
    if platform.system().lower() == "windows":
        try:
            import ctypes
            ctypes.windll.kernel32.SetConsoleOutputCP(65001)
        except Exception:
            pass
    app = QtWidgets.QApplication(sys.argv)
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()

